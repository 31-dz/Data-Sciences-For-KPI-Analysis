import pandas as pd
import numpy as np
from scipy import stats
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

def load_data(filename='lng_flare_data.csv'):
    """Load and prepare flare data"""
    df = pd.read_csv(filename)
    df['timestamp'] = pd.to_datetime(df['timestamp'])
    df['date'] = df['timestamp'].dt.date
    df['week'] = df['timestamp'].dt.isocalendar().week
    
    print(f"Data loaded successfully: {len(df):,} records")
    print(f"Date range: {df['timestamp'].min()} to {df['timestamp'].max()}")
    return df

def style_header(ws, row=1):
    """Apply header styling to a worksheet"""
    for cell in ws[row]:
        if cell.value:  # Only style cells with content
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

def descriptive_statistics(df):
    """Comprehensive descriptive statistics"""
    total_rate = df['total_flare_rate_m3_per_hour']
    
    stats_dict = {
        'Metric': ['Count', 'Mean', 'Median', 'Std Dev', 'Min', 'Max', 'Range',
                   'Q1 (25%)', 'Q3 (75%)', 'IQR', 'Skewness', 'Kurtosis',
                   '5th Percentile', '10th Percentile', '25th Percentile', 
                   '50th Percentile', '75th Percentile', '90th Percentile', 
                   '95th Percentile', '99th Percentile'],
        'Value': [
            len(total_rate),
            total_rate.mean(),
            total_rate.median(),
            total_rate.std(),
            total_rate.min(),
            total_rate.max(),
            total_rate.max() - total_rate.min(),
            total_rate.quantile(0.25),
            total_rate.quantile(0.75),
            total_rate.quantile(0.75) - total_rate.quantile(0.25),
            stats.skew(total_rate),
            stats.kurtosis(total_rate),
            total_rate.quantile(0.05),
            total_rate.quantile(0.10),
            total_rate.quantile(0.25),
            total_rate.quantile(0.50),
            total_rate.quantile(0.75),
            total_rate.quantile(0.90),
            total_rate.quantile(0.95),
            total_rate.quantile(0.99),
        ]
    }
    
    return pd.DataFrame(stats_dict)

def normality_tests_by_cause(df):
    """Test for normality of distribution for each cause"""
    cause_cols = {
        'normal_operations_m3_per_hour': 'Normal Operations',
        'process_upset_m3_per_hour': 'Process Upset',
        'equipment_maintenance_m3_per_hour': 'Equipment Maintenance',
        'startup_shutdown_m3_per_hour': 'Startup/Shutdown',
        'emergency_relief_m3_per_hour': 'Emergency Relief',
        'compressor_trip_m3_per_hour': 'Compressor Trip',
        'instrument_failure_m3_per_hour': 'Instrument Failure'
    }
    
    results = []
    
    for col, name in cause_cols.items():
        data = df[col]
        active_data = data[data > 0]  # Only test active periods
        
        if len(active_data) < 3:
            results.append({
                'Cause': name,
                'Sample Size': len(active_data),
                'Mean (m³/hr)': 0,
                'Std Dev': 0,
                'Shapiro-Wilk Statistic': 'N/A',
                'Shapiro-Wilk P-value': 'N/A',
                'Shapiro Result': 'Insufficient Data',
                'KS Statistic': 'N/A',
                'KS P-value': 'N/A',
                'KS Result': 'Insufficient Data',
                'Anderson Statistic': 'N/A',
                'Interpretation': 'Not enough data for testing'
            })
            continue
        
        # Shapiro-Wilk test (on sample if too large)
        sample_size = min(5000, len(active_data))
        sample = active_data.sample(n=sample_size, random_state=42) if len(active_data) > 5000 else active_data
        shapiro_stat, shapiro_p = stats.shapiro(sample)
        
        # Kolmogorov-Smirnov test
        ks_stat, ks_p = stats.kstest(active_data, 'norm', 
                                      args=(active_data.mean(), active_data.std()))
        
        # Anderson-Darling test
        anderson_result = stats.anderson(active_data, dist='norm')
        
        # Interpretation
        shapiro_normal = shapiro_p > 0.05
        ks_normal = ks_p > 0.05
        
        if shapiro_normal and ks_normal:
            interpretation = 'Data appears normally distributed'
        elif not shapiro_normal and not ks_normal:
            interpretation = 'Data is NOT normally distributed'
        else:
            interpretation = 'Mixed results - likely not normal'
        
        results.append({
            'Cause': name,
            'Sample Size': len(active_data),
            'Mean (m³/hr)': round(active_data.mean(), 2),
            'Std Dev': round(active_data.std(), 2),
            'Shapiro-Wilk Statistic': round(shapiro_stat, 6),
            'Shapiro-Wilk P-value': f"{shapiro_p:.6e}" if shapiro_p < 0.001 else round(shapiro_p, 6),
            'Shapiro Result': 'Normal' if shapiro_normal else 'Not Normal',
            'KS Statistic': round(ks_stat, 6),
            'KS P-value': f"{ks_p:.6e}" if ks_p < 0.001 else round(ks_p, 6),
            'KS Result': 'Normal' if ks_normal else 'Not Normal',
            'Anderson Statistic': round(anderson_result.statistic, 6),
            'Interpretation': interpretation
        })
    
    return pd.DataFrame(results)

def outlier_analysis(df):
    """Identify and analyze outliers"""
    total_rate = df['total_flare_rate_m3_per_hour']
    
    # IQR method
    Q1 = total_rate.quantile(0.25)
    Q3 = total_rate.quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    
    iqr_outliers = df[(total_rate < lower_bound) | (total_rate > upper_bound)]
    
    # Z-score method
    z_scores = np.abs(stats.zscore(total_rate))
    z_outliers = df[z_scores > 3]
    
    # Modified Z-score method
    median = total_rate.median()
    mad = np.median(np.abs(total_rate - median))
    modified_z_scores = 0.6745 * (total_rate - median) / mad
    modified_z_outliers = df[np.abs(modified_z_scores) > 3.5]
    
    summary = pd.DataFrame({
        'Method': ['IQR (1.5× IQR)', 'Z-Score (|z| > 3)', 'Modified Z-Score (MAD, |z| > 3.5)'],
        'Outliers Found': [len(iqr_outliers), len(z_outliers), len(modified_z_outliers)],
        'Percentage': [
            f"{len(iqr_outliers)/len(df)*100:.2f}%",
            f"{len(z_outliers)/len(df)*100:.2f}%",
            f"{len(modified_z_outliers)/len(df)*100:.2f}%"
        ],
        'Lower Bound': [f"{lower_bound:.2f}", 'N/A', 'N/A'],
        'Upper Bound': [f"{upper_bound:.2f}", 'N/A', 'N/A']
    })
    
    # Top outliers
    if len(iqr_outliers) > 0:
        top_outliers = iqr_outliers.nlargest(10, 'total_flare_rate_m3_per_hour')[
            ['timestamp', 'total_flare_rate_m3_per_hour', 'dominant_cause', 'severity']
        ].copy()
        top_outliers.columns = ['Timestamp', 'Flare Rate (m³/hr)', 'Dominant Cause', 'Severity']
    else:
        top_outliers = pd.DataFrame(columns=['Timestamp', 'Flare Rate (m³/hr)', 'Dominant Cause', 'Severity'])
    
    return summary, top_outliers

def cause_correlation_analysis(df):
    """Analyze correlations between different causes"""
    cause_cols = [
        'normal_operations_m3_per_hour',
        'process_upset_m3_per_hour',
        'equipment_maintenance_m3_per_hour',
        'startup_shutdown_m3_per_hour',
        'emergency_relief_m3_per_hour',
        'compressor_trip_m3_per_hour',
        'instrument_failure_m3_per_hour'
    ]
    
    corr_matrix = df[cause_cols].corr().round(4)
    
    # Rename columns for readability
    short_names = ['Normal Ops', 'Process Upset', 'Equip Maint', 
                   'Startup/SD', 'Emergency', 'Compressor', 'Instrument']
    corr_matrix.columns = short_names
    corr_matrix.index = short_names
    
    # Find strongest correlations
    strong_corr = []
    for i in range(len(cause_cols)):
        for j in range(i+1, len(cause_cols)):
            corr_val = df[cause_cols].corr().iloc[i, j]
            if abs(corr_val) > 0.05:
                strong_corr.append({
                    'Cause 1': short_names[i],
                    'Cause 2': short_names[j],
                    'Correlation': round(corr_val, 4)
                })
    
    if strong_corr:
        strong_corr_df = pd.DataFrame(strong_corr).sort_values('Correlation', 
                                                               key=abs, 
                                                               ascending=False)
    else:
        strong_corr_df = pd.DataFrame(columns=['Cause 1', 'Cause 2', 'Correlation'])
    
    return corr_matrix, strong_corr_df

def temporal_analysis(df):
    """Analyze temporal patterns"""
    # Hourly patterns
    hourly = df.groupby('hour')['total_flare_rate_m3_per_hour'].agg([
        'mean', 'std', 'min', 'max', 'count'
    ]).round(2)
    hourly.index.name = 'Hour'
    hourly.columns = ['Mean (m³/hr)', 'Std Dev', 'Min', 'Max', 'Count']
    
    # Day of week patterns
    dow_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    daily = df.groupby('day_of_week')['total_flare_rate_m3_per_hour'].agg([
        'mean', 'std', 'min', 'max', 'count'
    ]).reindex(dow_order).round(2)
    daily.index.name = 'Day of Week'
    daily.columns = ['Mean (m³/hr)', 'Std Dev', 'Min', 'Max', 'Count']
    
    # Monthly patterns
    month_order = ['January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    monthly = df.groupby('month')['total_flare_rate_m3_per_hour'].agg([
        'mean', 'sum', 'std', 'min', 'max', 'count'
    ]).reindex(month_order).round(2)
    monthly.index.name = 'Month'
    monthly.columns = ['Mean (m³/hr)', 'Total (m³)', 'Std Dev', 'Min', 'Max', 'Count']
    
    # Weekly aggregates
    weekly = df.groupby('week')['total_flare_rate_m3_per_hour'].sum().round(2)
    weekly_df = pd.DataFrame({
        'Week': weekly.index,
        'Total Flare (m³)': weekly.values
    })
    
    return hourly, daily, monthly, weekly_df

def severity_analysis(df):
    """Analyze severity patterns"""
    severity_stats = df.groupby('severity').agg({
        'total_flare_rate_m3_per_hour': ['count', 'mean', 'std', 'min', 'max', 'sum']
    }).round(2)
    
    severity_stats.columns = ['Count', 'Mean (m³/hr)', 'Std Dev', 'Min', 'Max', 'Total (m³)']
    severity_stats['Percentage'] = (severity_stats['Count'] / len(df) * 100).round(2)
    severity_stats = severity_stats[['Count', 'Percentage', 'Mean (m³/hr)', 'Std Dev', 'Min', 'Max', 'Total (m³)']]
    
    return severity_stats

def cause_specific_statistics(df):
    """Detailed statistics for each cause"""
    cause_cols = {
        'normal_operations_m3_per_hour': 'Normal Operations',
        'process_upset_m3_per_hour': 'Process Upset',
        'equipment_maintenance_m3_per_hour': 'Equipment Maintenance',
        'startup_shutdown_m3_per_hour': 'Startup/Shutdown',
        'emergency_relief_m3_per_hour': 'Emergency Relief',
        'compressor_trip_m3_per_hour': 'Compressor Trip',
        'instrument_failure_m3_per_hour': 'Instrument Failure'
    }
    
    results = []
    for col, name in cause_cols.items():
        data = df[col]
        active = data[data > 0]
        
        results.append({
            'Cause': name,
            'Total Volume (m³)': round(data.sum(), 2),
            'Active Hours': len(active),
            'Active %': f"{len(active)/len(df)*100:.2f}%",
            'Mean Active (m³/hr)': round(active.mean(), 2) if len(active) > 0 else 0,
            'Std Dev Active': round(active.std(), 2) if len(active) > 0 else 0,
            'Max Rate (m³/hr)': round(active.max(), 2) if len(active) > 0 else 0
        })
    
    return pd.DataFrame(results)

def trend_analysis(df):
    """Analyze long-term trends"""
    # Calculate daily totals for trend
    daily = df.groupby('date')['total_flare_rate_m3_per_hour'].sum().reset_index()
    daily['day_number'] = range(len(daily))
    
    # Linear regression
    slope, intercept, r_value, p_value, std_err = stats.linregress(
        daily['day_number'], daily['total_flare_rate_m3_per_hour']
    )
    
    trend_summary = pd.DataFrame({
        'Metric': ['Slope (m³/day per day)', 'R-squared', 'P-value', 'Std Error', 
                   'Interpretation'],
        'Value': [
            round(slope, 4),
            round(r_value**2, 4),
            f"{p_value:.6e}" if p_value < 0.001 else round(p_value, 6),
            round(std_err, 4),
            f"{'Significant increasing' if p_value < 0.05 and slope > 0 else 'Significant decreasing' if p_value < 0.05 and slope < 0 else 'No significant'} trend"
        ]
    })
    
    # Moving averages
    df_sorted = df.sort_values('timestamp').copy()
    df_sorted['MA_24h'] = df_sorted['total_flare_rate_m3_per_hour'].rolling(window=24).mean()
    df_sorted['MA_168h'] = df_sorted['total_flare_rate_m3_per_hour'].rolling(window=168).mean()
    
    ma_summary = pd.DataFrame({
        'Moving Average': ['24-hour MA', '7-day MA (168h)'],
        'Min (m³/hr)': [
            round(df_sorted['MA_24h'].min(), 2),
            round(df_sorted['MA_168h'].min(), 2)
        ],
        'Max (m³/hr)': [
            round(df_sorted['MA_24h'].max(), 2),
            round(df_sorted['MA_168h'].max(), 2)
        ],
        'Mean (m³/hr)': [
            round(df_sorted['MA_24h'].mean(), 2),
            round(df_sorted['MA_168h'].mean(), 2)
        ]
    })
    
    return trend_summary, ma_summary

def dominant_cause_analysis(df):
    """Analyze dominant cause distribution"""
    dominant_counts = df['dominant_cause'].value_counts().reset_index()
    dominant_counts.columns = ['Cause', 'Frequency']
    dominant_counts['Percentage'] = (dominant_counts['Frequency'] / len(df) * 100).round(2)
    
    # Rename causes for readability
    cause_names = {
        'normal_operations': 'Normal Operations',
        'process_upset': 'Process Upset',
        'equipment_maintenance': 'Equipment Maintenance',
        'startup_shutdown': 'Startup/Shutdown',
        'emergency_relief': 'Emergency Relief',
        'compressor_trip': 'Compressor Trip',
        'instrument_failure': 'Instrument Failure'
    }
    dominant_counts['Cause'] = dominant_counts['Cause'].map(cause_names)
    
    return dominant_counts

def main():
    """Run complete statistical analysis and export to Excel"""
    print("\n" + "="*80)
    print("LNG FLARE GAS STATISTICAL ANALYSIS")
    print("="*80 + "\n")
    
    print("Loading data...")
    df = load_data()
    
    print("\nRunning analyses...")
    
    # Create Excel writer
    output_file = 'lng_flare_statistical_analysis.xlsx'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # 1. Summary
        print("  - Creating summary...")
        summary_data = pd.DataFrame({
            'Metric': ['Total Records', 'Date Range', 'Total Flare Gas (m³)', 
                      'Average Rate (m³/hr)', 'Peak Rate (m³/hr)', 'Generated'],
            'Value': [
                len(df),
                f"{df['timestamp'].min()} to {df['timestamp'].max()}",
                f"{df['total_flare_rate_m3_per_hour'].sum():,.2f}",
                f"{df['total_flare_rate_m3_per_hour'].mean():,.2f}",
                f"{df['total_flare_rate_m3_per_hour'].max():,.2f}",
                pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        })
        summary_data.to_excel(writer, sheet_name='Summary', index=False)
        
        # 2. Descriptive Statistics
        print("  - Calculating descriptive statistics...")
        desc_stats = descriptive_statistics(df)
        desc_stats.to_excel(writer, sheet_name='Descriptive Stats', index=False)
        
        # 3. Normality Tests (by cause)
        print("  - Running normality tests by cause...")
        normality_df = normality_tests_by_cause(df)
        normality_df.to_excel(writer, sheet_name='Normality Tests', index=False)
        
        # 4. Outlier Analysis
        print("  - Analyzing outliers...")
        outlier_summary, top_outliers = outlier_analysis(df)
        outlier_summary.to_excel(writer, sheet_name='Outliers', index=False, startrow=0)
        if not top_outliers.empty:
            top_outliers.to_excel(writer, sheet_name='Outliers', index=False, startrow=len(outlier_summary)+3)
        
        # 5. Correlation Analysis
        print("  - Calculating correlations...")
        corr_matrix, strong_corr = cause_correlation_analysis(df)
        corr_matrix.to_excel(writer, sheet_name='Correlations', startrow=0)
        if not strong_corr.empty:
            strong_corr.to_excel(writer, sheet_name='Correlations', index=False, startrow=len(corr_matrix)+3)
        
        # 6. Temporal Analysis
        print("  - Analyzing temporal patterns...")
        hourly, daily, monthly, weekly = temporal_analysis(df)
        hourly.to_excel(writer, sheet_name='Temporal-Hourly')
        daily.to_excel(writer, sheet_name='Temporal-Daily')
        monthly.to_excel(writer, sheet_name='Temporal-Monthly')
        weekly.to_excel(writer, sheet_name='Temporal-Weekly', index=False)
        
        # 7. Severity Analysis
        print("  - Analyzing severity levels...")
        severity = severity_analysis(df)
        severity.to_excel(writer, sheet_name='Severity Analysis')
        
        # 8. Cause-Specific Statistics
        print("  - Calculating cause-specific statistics...")
        cause_stats = cause_specific_statistics(df)
        cause_stats.to_excel(writer, sheet_name='Cause Statistics', index=False)
        
        # 9. Dominant Cause Analysis
        print("  - Analyzing dominant causes...")
        dominant = dominant_cause_analysis(df)
        dominant.to_excel(writer, sheet_name='Dominant Causes', index=False)
        
        # 10. Trend Analysis
        print("  - Analyzing trends...")
        trend_sum, ma_sum = trend_analysis(df)
        trend_sum.to_excel(writer, sheet_name='Trend Analysis', index=False, startrow=0)
        ma_sum.to_excel(writer, sheet_name='Trend Analysis', index=False, startrow=len(trend_sum)+3)
        
        print("\nApplying formatting...")
        
        # Apply formatting to all sheets
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            style_header(ws)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    print(f"\n{'='*80}")
    print(f"✓ Analysis complete! Results saved to: {output_file}")
    print(f"{'='*80}\n")
    
    print("Worksheets created:")
    worksheets = [
        'Summary', 'Descriptive Stats', 'Normality Tests (by cause)',
        'Outliers', 'Correlations', 'Temporal-Hourly', 'Temporal-Daily',
        'Temporal-Monthly', 'Temporal-Weekly', 'Severity Analysis',
        'Cause Statistics', 'Dominant Causes', 'Trend Analysis'
    ]
    for i, ws in enumerate(worksheets, 1):
        print(f"  {i:2d}. {ws}")
    
    print("\n" + "="*80 + "\n")

if __name__ == "__main__":
    main()